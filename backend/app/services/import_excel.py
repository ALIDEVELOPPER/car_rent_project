"""Import de la flotte et des clients depuis un fichier Excel (.xlsx).

Toute agence a déjà ses véhicules et ses clients dans un tableur — l'import
supprime la principale friction à l'adoption. On accepte des en-têtes de
colonnes en français, insensibles à la casse et aux accents.
"""
import io
import unicodedata
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.models import Client, Vehicule
from app.models.enums import Carburant, TypePieceIdentite


def _norm(value) -> str:
    if value is None:
        return ""
    txt = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode()
    return txt.strip().lower()


# Colonnes reconnues -> attribut du modèle. La 1re clé est celle du modèle généré.
VEHICULE_COLS = {
    "marque": ["marque"],
    "modele": ["modele", "model"],
    "immatriculation": ["immatriculation", "matricule", "plaque"],
    "categorie": ["categorie", "type"],
    "tarif_jour": ["tarif/jour", "tarif jour", "tarif journalier", "prix/jour", "prix jour"],
    "annee": ["annee", "année"],
    "couleur": ["couleur"],
    "kilometrage": ["kilometrage", "km"],
    "carburant": ["carburant"],
}
VEHICULE_REQUIS = ("marque", "modele", "immatriculation", "categorie", "tarif_jour")

CLIENT_COLS = {
    "nom": ["nom"],
    "prenom": ["prenom"],
    "telephone": ["telephone", "tel", "gsm"],
    "email": ["email", "e-mail", "mail"],
    "adresse": ["adresse"],
    "numero_piece_identite": ["cin", "n piece", "numero piece", "piece identite", "n cin"],
    "numero_permis": ["permis", "n permis", "numero permis"],
}
CLIENT_REQUIS = ("nom", "prenom", "telephone")

_CARBURANTS = {_norm(c.value): c for c in Carburant}
_CARBURANTS.update({"gasoil": Carburant.DIESEL, "gazole": Carburant.DIESEL, "electric": Carburant.ELECTRIQUE})


def _map_headers(row, col_defs: dict) -> dict:
    """{index_colonne: nom_champ} d'après la ligne d'en-tête."""
    mapping = {}
    for idx, cell in enumerate(row):
        h = _norm(cell)
        if not h:
            continue
        for champ, alias in col_defs.items():
            if h in alias or h == champ:
                mapping[idx] = champ
    return mapping


def _read_rows(file_bytes: bytes, col_defs: dict):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return [], {}
    mapping = _map_headers(header, col_defs)
    lignes = []
    for raw in rows:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        item = {}
        for idx, champ in mapping.items():
            if idx < len(raw):
                val = raw[idx]
                item[champ] = str(val).strip() if val is not None else ""
        lignes.append(item)
    return lignes, mapping


def _decimal(value) -> Decimal:
    return Decimal(str(value).replace(",", ".").replace(" ", ""))


def import_vehicules(file_bytes: bytes) -> dict:
    lignes, mapping = _read_rows(file_bytes, VEHICULE_COLS)
    manquants = [c for c in VEHICULE_REQUIS if c not in mapping.values()]
    if manquants:
        return {"error": f"Colonnes obligatoires absentes : {', '.join(manquants)}"}

    existants = {
        _norm(v) for v in db.session.execute(db.select(Vehicule.immatriculation)).scalars()
    }
    cree, ignore, erreurs = 0, 0, []
    for i, item in enumerate(lignes, start=2):
        if any(not item.get(c) for c in VEHICULE_REQUIS):
            erreurs.append(f"Ligne {i} : champ obligatoire vide")
            continue
        immat = item["immatriculation"]
        if _norm(immat) in existants:
            ignore += 1
            continue
        try:
            tarif = _decimal(item["tarif_jour"])
            if tarif <= 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            erreurs.append(f"Ligne {i} : tarif/jour invalide ({item['tarif_jour']!r})")
            continue

        v = Vehicule(
            marque=item["marque"], modele=item["modele"], immatriculation=immat,
            categorie=item["categorie"], tarif_jour=tarif,
        )
        if item.get("annee"):
            try:
                v.annee = int(float(item["annee"]))
            except ValueError:
                pass
        if item.get("kilometrage"):
            try:
                v.kilometrage = int(float(str(item["kilometrage"]).replace(" ", "")))
            except ValueError:
                pass
        v.couleur = item.get("couleur") or None
        if item.get("carburant"):
            v.carburant = _CARBURANTS.get(_norm(item["carburant"]))

        db.session.add(v)
        existants.add(_norm(immat))
        cree += 1

    db.session.commit()
    return {"cree": cree, "ignore": ignore, "erreurs": erreurs}


def import_clients(file_bytes: bytes) -> dict:
    lignes, mapping = _read_rows(file_bytes, CLIENT_COLS)
    manquants = [c for c in CLIENT_REQUIS if c not in mapping.values()]
    if manquants:
        return {"error": f"Colonnes obligatoires absentes : {', '.join(manquants)}"}

    cree, ignore, erreurs = 0, 0, []
    for i, item in enumerate(lignes, start=2):
        if any(not item.get(c) for c in CLIENT_REQUIS):
            erreurs.append(f"Ligne {i} : champ obligatoire vide")
            continue
        c = Client(
            nom=item["nom"], prenom=item["prenom"], telephone=item["telephone"],
            email=item.get("email") or None, adresse=item.get("adresse") or None,
            numero_piece_identite=item.get("numero_piece_identite") or None,
            numero_permis=item.get("numero_permis") or None,
        )
        if c.numero_piece_identite:
            c.type_piece_identite = TypePieceIdentite.CNI
        db.session.add(c)
        cree += 1

    db.session.commit()
    return {"cree": cree, "ignore": ignore, "erreurs": erreurs}


def modele_vehicules_xlsx() -> bytes:
    return _modele(["Marque", "Modèle", "Immatriculation", "Catégorie", "Tarif/jour",
                    "Année", "Couleur", "Kilométrage", "Carburant"],
                   ["Dacia", "Logan", "1234-A-56", "Berline", "250", "2021", "Gris", "45000", "diesel"])


def modele_clients_xlsx() -> bytes:
    return _modele(["Nom", "Prénom", "Téléphone", "Email", "Adresse", "CIN", "Permis"],
                   ["Bennani", "Yassine", "0661234567", "y.bennani@example.ma",
                    "12 rue X, Casablanca", "BE123456", "12/345678"])


def _modele(entetes: list[str], exemple: list[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(entetes)
    ws.append(exemple)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
